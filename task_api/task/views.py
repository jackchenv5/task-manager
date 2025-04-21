import io
import math
import string
from datetime import datetime

from django.contrib.auth import get_user_model
from django.contrib.auth.models import User
from django.http import HttpResponse
from rest_framework.views import APIView

from common.utils import parse_date_range, is_date_range, workdays_between_with_holidays

User = get_user_model()
from rest_framework import viewsets
from user.models import Group
from task.models import TaskStatus, TaskCategory, Task
from task.serializers import TaskStatusSerializer, TaskCategorySerializer, TaskSerializer

from rest_framework.pagination import PageNumberPagination
from rest_framework.response import Response
from rest_framework import status
from django_filters import rest_framework as filters
from common.email import send_email, generate_email_body
import common
import pandas as pd

from openpyxl import Workbook

from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

from django.http import FileResponse
from django.core.files.storage import default_storage
from django.conf import settings
import os


# 自定义分页类（可选，如果需要自定义每页显示的条目数）
class StandardPagination(PageNumberPagination):
    page_size_query_param = 'pageSize'  # 允许客户端通过查询参数来覆盖默认设置


class TaskStatusViewSet(viewsets.ModelViewSet):
    """
    This viewset automatically provides `list` and `retrieve` actions.hello
    """
    queryset = TaskStatus.objects.all()
    serializer_class = TaskStatusSerializer


class TaskCategoryViewSet(viewsets.ModelViewSet):
    """
    This viewset automatically provides `list` and `retrieve` actions.hello
    """
    queryset = TaskCategory.objects.all()
    serializer_class = TaskCategorySerializer


class TaskFilter(filters.FilterSet):
    # 假设您的时间字段是datetime类型，使用DateTimeFilter
    start_time = filters.DateTimeFilter(field_name="start_time", lookup_expr='gte')
    deadline_time = filters.DateTimeFilter(field_name="deadline_time", lookup_expr='lte')
    receiver = filters.NumberFilter(field_name="receiver_id")  # 假设receiver是一个关联到User的外键
    creator = filters.NumberFilter(field_name="creator_id")  # 假设receiver是一个关联到User的外键
    status = filters.NumberFilter(field_name="status_id")  # 假设status是一个关联到Status的外键

    # 使用CharFilter进行文本搜索，并使用icontains进行不区分大小写的包含搜索
    search_text = filters.CharFilter(
        method='filter_search_text',
        label='Search text in content, challenge, feedback, name'
    )
    flag_time = filters.DateTimeFilter(method='filter_flag_time',)

    class Meta:
        model = Task
        fields = ['start_time', 'deadline_time', 'receiver', 'creator', 'status', 'project']  # 这里只列出用于自动生成查询参数的字段

    def filter_search_text(self, queryset, name, value):
        from django.db.models import Q
        if value and name == 'search_text':
            return queryset.filter(
                Q(id=int(value) if value.isdigit() else -1) |
                Q(content__icontains=value) |
                Q(challenge__icontains=value) |
                Q(feedback__icontains=value) |
                Q(project__icontains=value) |
                Q(name__icontains=value)
            )

    def filter_flag_time(self, queryset, name, value):
        if value and name == "flag_time":
            return queryset.filter(start_time__lte = value,deadline_time__gte = value)

class TaskViewSet(viewsets.ModelViewSet):
    """
    This viewset automatically provides `list` and `retrieve` actions.hello
    """
    queryset = Task.objects.all()
    serializer_class = TaskSerializer

    def update(self, request, *args, **kwargs):
        partial = True  # 允许部分更新
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        return Response(serializer.data)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        sub_tasks = instance.task_set.all()
        add_emails = instance.receiver.get_group_leader_email()
        if instance.creator:
            add_emails.append(instance.creator.email)
        cc_emails = instance.get_sender_email()
        add_content = ""
        if sub_tasks.count() > 0:
            # 有子任务，返回子任务
            for tmp_task in sub_tasks:
                publisher = tmp_task.publisher.email if tmp_task.publisher else ""
                receiver = tmp_task.receiver.email if tmp_task.receiver else ""
                creator = tmp_task.creator.email if tmp_task.creator else ""
                senders = tmp_task.get_sender_email()
                if senders:
                    cc_emails = list(set(cc_emails + senders))
                if publisher and not publisher in add_emails:
                    add_emails.append(publisher)
                if receiver and not receiver not in add_emails:
                    add_emails.append(receiver)
                if creator and not add_emails in add_emails:
                    add_emails.append(creator)
            extra_str = '<p style="font-size:12px;"><b>由于关联任务删除，上述关联任务设置为空！</b></p>'
            add_content = generate_email_body(list(sub_tasks), status="子任务", add_content=extra_str, has_footer=False)
        email_body = generate_email_body(instance, status="被删除", add_content=add_content)
        if instance.status.name == common.TASK_STATUS_PROGRESS:
            if instance.publisher and instance.publisher.email not in add_emails:
                add_emails.append(instance.publisher.email)
            if instance.receiver and instance.receiver.email not in add_emails:
                add_emails.append(instance.receiver.email)
            if instance.sender:
                instance_senders = instance.get_sender_email()
                cc_emails = list(set(cc_emails + instance_senders))
            send_email(email_body, list(set(add_emails)), cc=cc_emails, subject=instance.name.replace("\n",""), action="任务删除")
        else:
            send_email(email_body, list(set(add_emails)), cc=[], subject=instance.name.replace("\n",""), action="任务删除")
        self.perform_destroy(instance)
        return Response(status=status.HTTP_204_NO_CONTENT)

    def list(self, request, *args, **kwargs):
        # 获取查询参数  
        filter_params = request.query_params.dict()
        workload_intensity_order = False
        # 从查询参数中提取group（如果存在）  
        group_id = filter_params.pop('group', None)
        order_filed = filter_params.pop('field', None)
        if order_filed == 'receiver_name':
            order_filed = 'receiver'
        if order_filed == 'workload_intensity':
            order_filed = ''
            workload_intensity_order = True
        order = filter_params.pop('order', None)

        # 创建FilterSet实例
        task_filter = TaskFilter(data=filter_params, queryset=Task.objects.filter(workload__gt=0))

        # 检查过滤是否有效  
        if not task_filter.is_valid():
            return Response(task_filter.errors, status=status.HTTP_400_BAD_REQUEST)

        filtered_tasks = task_filter.qs

        # 如果提供了group_id，则进一步过滤tasks  
        if group_id:
            try:
                group = Group.objects.get(id=group_id)
                # 使用Django ORM来查询接收者在给定组内的任务  
                filtered_tasks = filtered_tasks.filter(receiver__group__id=group_id)
            except Group.DoesNotExist:
                # 处理组不存在的情况
                return Response({"error": "Group not found"}, status=status.HTTP_404_NOT_FOUND)
                # 排序 id降序
        if order_filed:
            if order == 'descend':
                filtered_tasks = filtered_tasks.order_by(f'-{order_filed}')
            else:
                filtered_tasks = filtered_tasks.order_by(order_filed)
        else:
            filtered_tasks = filtered_tasks.order_by('-pk')
        # 使用TaskSerializer序列化查询集  
        data = TaskSerializer(filtered_tasks, many=True).data
        # workload_intensity
        if workload_intensity_order:
            if order == 'descend':
                data = sorted(data, key=lambda x: x['workload_intensity'], reverse=True)
            else:
                data = sorted(data, key=lambda x: x['workload_intensity'])
        # 计算
        start_min, end_max = None, None
        task_user_info = {}
        for tmp in data:
            user = tmp['receiver']
            if not tmp['status_name'] in ["进行中", "待下发"]:
                continue
            start_time = datetime.strptime(tmp['start_time'], '%Y-%m-%d')
            end_time = datetime.strptime(tmp['deadline_time'], '%Y-%m-%d')
            if not start_min or start_time < start_min:
                start_min = start_time
            if not end_max or end_time > end_max:
                end_max = end_time
            if user not in task_user_info:
                task_user_info[user] = 0.0
            task_user_info[user] += tmp['workload']
        work_days = workdays_between_with_holidays(start_min, end_max)
        ret_data = []
        for tmp in data:
            user = tmp['receiver']
            tmp['duration_workload'] = ""
            if tmp['status_name'] in ["进行中", "待下发"]:
                if work_days > 0 and user in task_user_info:
                    tmp['duration_workload'] = "%s|%s|%s" % (int(100 * task_user_info[user] / work_days), task_user_info[user], work_days)
            ret_data.append(tmp)
        # 返回序列化后的数据  
        # return Response(serializer.data)
        return Response({
            'code': 0,
            'message': 'ok',
            'result': {
                'items': ret_data,
                'total': filtered_tasks.count(),  # 这里获取总条目数
            },
            'type': 'success',
        })


class NotifyTasksByReceiverAPI(APIView):
    def get(self, request, *args, **kwargs):
        # 假设我们从请求体中接收一个包含ids的列表
        task_ids_str = request.query_params.get('ids', '')
        publisher_id = request.query_params.get('publisher', '')
        publisher = User.objects.get(id=publisher_id)
        task_ids = list(map(int, task_ids_str.split(','))) if task_ids_str else []
        if not isinstance(task_ids, list):
            return Response({"error": "IDs not get!"}, status=400)

        # 过滤出需要通知的任务
        tasks_to_notify = Task.objects.filter(id__in=task_ids, status_id=1)

        # 对每个不同的receiver进行分组，并发送邮件
        receivers_tasks = {}
        for task in tasks_to_notify:
            if task.receiver_id not in receivers_tasks:
                receivers_tasks[task.receiver_id] = []
            receivers_tasks[task.receiver_id].append(task)

        # 通知 组长 TL 本人
        for receiver_id, tasks in receivers_tasks.items():
            receiver = User.objects.get(id=receiver_id)  # 假设User是你的用户模型
            to = receiver.get_group_leader_email()
            if receiver.email not in to:
                to.append(receiver.email)
            if publisher.email not in to:
                to.append(publisher.email)
            # 抄送
            cc = []
            for task in tasks:
                # 任务TL
                if task.creator and task.creator.email not in to:
                    to.append(task.creator.email)
                # 抄送人
                tmp_cc = task.get_sender_email()
                if tmp_cc:
                    cc = list(set(cc + tmp_cc))
            email_body = generate_email_body(tasks, publisher.username)
            send_email(email_body, to, cc=cc, subject=receiver.username)

        # 跟新状态
        progress_status, _ = TaskStatus.objects.get_or_create(name=common.TASK_STATUS_PROGRESS)
        tasks_to_notify.update(status=progress_status, publisher=publisher)

        return Response({"status": "Emails sent successfully"}, status=200)


class ExportExcel(APIView):
    def get(self, request, *args, **kwargs):
        # 假设你有一个函数来获取任务数据
        # 获取查询参数
        filter_params = request.query_params.dict()
        # 从查询参数中提取group（如果存在）
        group_id = filter_params.pop('group', None)

        # 创建FilterSet实例
        task_filter = TaskFilter(data=filter_params, queryset=Task.objects.all())

        # 检查过滤是否有效
        if not task_filter.is_valid():
            return Response(task_filter.errors, status=status.HTTP_400_BAD_REQUEST)

        filtered_tasks = task_filter.qs

        # 如果提供了group_id，则进一步过滤tasks
        if group_id:
            try:
                group = Group.objects.get(id=group_id)
                # 使用Django ORM来查询接收者在给定组内的任务
                filtered_tasks = filtered_tasks.filter(receiver__group__id=group_id)
            except Group.DoesNotExist:
                # 处理组不存在的情况
                return Response({"error": "Group not found"}, status=status.HTTP_404_NOT_FOUND)
                # 排序 id降序
        if filtered_tasks.count() == 0:
            return Response({"error": "not tasks"}, status=status.HTTP_404_NOT_FOUND)
        # 准备数据框
        data = []
        time_range = {'min': "", 'max': ""}
        for task in filtered_tasks:
            if task.start_time and (not time_range['min'] or task.start_time < time_range['min']):
                time_range['min'] = task.start_time
            if task.start_time and (not time_range['max'] or task.deadline_time > time_range['max']):
                time_range['max'] = task.deadline_time
            row = {
                'ID': task.id,
                '任务名': task.name,
                '执行人': task.receiver.username if task.receiver else "未指定",  # 假设 TaskCategory 有一个 name 字段
                '发布人': task.publisher.username if task.publisher else "未指定",  # 假设 TaskCategory 有一个 name 字段
                '创建人': task.creator.username if task.creator else "未指定",  # 假设 TaskCategory 有一个 name 字段
                '开始时间': task.start_time.strftime("%Y-%m-%d") if task.start_time else "未指定",
                '截止时间': task.deadline_time.strftime("%Y-%m-%d") if task.deadline_time else "未指定",
                '工作量/天': task.workload,
                '实际工作量/天': task.act_workload,
                '项目': task.project,
                '关联任务': task.related_task.name if task.related_task else "未指定",
                '当前状态': task.status.name,
                '任务内容': task.content,
                '挑战目标': task.challenge,
                '反馈信息': task.feedback,
            }
            data.append(row)

        df = pd.DataFrame(data)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        if isinstance(time_range['min'], datetime):
            time_range['min'] = time_range['min'].strftime("%Y-%m-%d")
        if isinstance(time_range['max'], datetime):
            time_range['max'] = time_range['max'].strftime("%Y-%m-%d")
        excel_name = "tasks%s~%s.xlsx" % (time_range['min'], time_range['max'])
        print(excel_name)
        # content_disposition = 'attachment; filename="%s"'% excel_name
        content_disposition = 'attachment; filename="%s"' % excel_name
        print(content_disposition)
        response['Content-Disposition'] = content_disposition

        with pd.ExcelWriter(response, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)

        return response

class ReportTask(APIView):
    def get(self, request, *args, **kwargs):
        # 假设你有一个函数来获取任务数据
        # 获取查
        filter_params = request.query_params.dict()

        type = filter_params.get('type','个人报告')
        user = filter_params.get('user','')
        if type == '个人报告':
            scripts_path = os.path.join(settings.BASE_DIR, 'scripts/send_task_report_receiver.py')
        elif type == '组报告':
            scripts_path = os.path.join(settings.BASE_DIR, 'scripts/send_task_report_group.py --user %s'% user)
        else :
            scripts_path = os.path.join(settings.BASE_DIR, 'scripts/send_task_report_tl.py')
        os.system("/WebServer/cc/task-admin/task_api/env/bin/python %s"%scripts_path)
        # 从查询参数中提取group（如果存在）
        return Response({"status": "success"}, status=200)



class ExportTestExcel(APIView):
    def get_col_letter(cls, index):
        uppercase_letters = list(string.ascii_uppercase)
        new_index = index % len(uppercase_letters)
        count = int(index / len(uppercase_letters))
        COL_LETTER = uppercase_letters[new_index]
        if count > 0:
            AFFIX = uppercase_letters[count - 1]
            COL_LETTER = AFFIX + COL_LETTER
        return COL_LETTER

    def get_data(cls, tasks):
        date_range_list = []
        data_dict = {}
        for task in tasks:
            start_date = task['start_time'].strftime("%Y-%m-%d")
            deadline_date = task['deadline_time'].strftime("%Y-%m-%d")
            task['start_time'] = start_date
            task['deadline_time'] = deadline_date
            project_name = task['project']
            receiver__username = task['receiver__username']
            date_range = (start_date, deadline_date)
            if date_range not in date_range_list:
                date_range_list.append(date_range)
            if project_name not in data_dict:
                data_dict[project_name] = {}
            if receiver__username not in data_dict[project_name]:
                data_dict[project_name][receiver__username] = {}
            if date_range not in data_dict[project_name][receiver__username]:
                data_dict[project_name][receiver__username][date_range] = []
            del task['project']
            del task['receiver__username']
            del task['start_time']
            del task['deadline_time']
            data_dict[project_name][receiver__username][date_range].append(task)
        return data_dict, date_range_list

    def update_date_range_index(cls, row_num, ret_dict):
        for key, data in ret_dict.items():
            if key in ['project', 'receiver__username']:
                continue
            else:
                ret_dict[key]['cur_row_index'] += row_num

    def get_table_info(cls, data_list):
        ret_dict = {
            'project': {'col_index': 1, 'cur_row_index': 2, 'name': '项目', 'width': 30},
            'receiver__username': {'col_index': 2, 'cur_row_index': 2, 'name': '执行者', 'width': 10},
            # ('2024-08-01', '2024-08-14'): {'col_index': 3, 'cur_row': 2, 'sub': {'id': 3, 'name': 4, 'content': 5, 'challenge': 6, 'workload': 7}},
            # ('2024-08-15', '2024-08-29'): {'col_index': 8, 'cur_row': 2, 'sub': {'id': 8, 'name':9, 'content': 10, 'challenge': 11, 'workload': 12}}
        }
        width_config = {'A': 50, 'B': 10}
        header_row_1 = ['项目', '执行人']  # 通过函数获取
        header_row_2 = ['', '']  # 通过函数获取
        merge_list = ['A1:A2', 'B1:B2']

        template_dict = {'id': 3, 'name': 4, 'content': 5, 'challenge': 6, 'workload': 7}
        template_width_dict = {'id': 10, 'name': 30, 'content': 50, 'challenge': 50, 'workload': 10}
        template_dict_len = len(template_dict)
        name_list = ['ID', '任务名', '任务内容', '任务目标', '工作量']
        template_dict_list = ['' for i in range(template_dict_len - 1)]
        col_index_base = 3
        cur_row_index_base = 2
        len_dict = len(template_dict)
        sorted_dates = sorted(data_list, key=lambda d: datetime.strptime(d[0], "%Y-%m-%d"))  # key参数实际上是可选的，因为datetime对象已经可比较
        date_range_count = 0
        letter_index = 2
        for date_range in sorted_dates:
            header_row_1 += ['%s-%s' % (date_range[0], date_range[1])] + template_dict_list
            start_letter = cls.get_col_letter(len(header_row_2))
            end_letter = cls.get_col_letter(len(header_row_2) + template_dict_len - 1)
            merge_list.append('%s1:%s1' % (start_letter, end_letter))
            header_row_2 += name_list
            ret_dict[date_range] = {'col_index': -1, 'cur_row_index': -1, 'sub': {}}
            ret_dict[date_range]['col_index'] = col_index_base + len_dict * date_range_count
            ret_dict[date_range]['cur_row_index'] = cur_row_index_base
            for key, value in template_dict.items():
                width_config[cls.get_col_letter(letter_index)] = template_width_dict[key]
                ret_dict[date_range]['sub'][key] = template_dict[key] + len_dict * date_range_count
                letter_index += 1
            date_range_count += 1
        print(width_config)
        return ret_dict, header_row_1, header_row_2, merge_list, width_config

    def create_excel_with_data(cls, tasks):
        # 创建一个新的Excel工作簿
        wb = Workbook()
        ws = wb.active

        real_data, date_range_list = cls.get_data(tasks)
        real_table_info, header_row_1, header_row_2, merge_list, width_config = cls.get_table_info(date_range_list)
        for key, value in width_config.items():
            ws.column_dimensions[key].width = value
        # 项目最长的排前面
        # header_row_1 = ['项目', '执行人', '2024-08-01~2024-08-14', '', '', '', '', '2024-08-15~2024-08-21', '', '', '', '']  # 通过函数获取
        # header_row_2 = ['    ', '     ', 'ID', '任务名', '任务内容', '任务目标', '工作量', 'ID', '任务名', '任务内容', '任务目标', '工作量']  # 通过函数获取

        # 写入表头
        # 字体样式：加粗、大小为12
        bold_font = Font(bold=True, size=14, color="FFFFFF")
        font2 = Font(bold=True, size=12, color="FFFFFF")
        # 填充样式：浅灰色背景
        fill1 = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
        fill2 = PatternFill(start_color='666666', end_color='666666', fill_type='solid')

        # 边框样式：细边框
        thin_border = Side(border_style="thin", color="000000")
        border = Border(top=thin_border, left=thin_border, right=thin_border, bottom=thin_border)
        header_alignment = Alignment(horizontal='center', vertical='center')  # 水平和垂直居中

        for col_num, header in enumerate(header_row_1, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = bold_font
            cell.fill = fill1
            cell.border = border
            cell.alignment = header_alignment
        for col_num, header in enumerate(header_row_2, 1):
            cell = ws.cell(row=2, column=col_num, value=header)
            cell.font = font2
            cell.fill = fill2
            cell.border = border
            cell.alignment = header_alignment
        # 获取用户在各个时间段的任务数，获取最大任务 project_user_max_len
        for project, data in real_data.items():
            max_info = cls.get_project_max(data)
            project_max = max_info['project_max']
            # 填充项目列
            cur_project_row_index = real_table_info['project']['cur_row_index']
            for project_index in range(project_max):
                cur_project_row_index += 1
                # 是否需要合并
                # if row_index == 1:
                #     ws.cell(row=row_index, column=table_info['project']['col_index'],value=project)
                ws.cell(row=cur_project_row_index, column=real_table_info['project']['col_index'], value=project)
            # 合并项目列
            # 填充用户

            for user, date_range_data in data.items():
                user_max = max_info['users'][user]
                tmp_user_row_index = real_table_info['receiver__username']['cur_row_index']

                for user_index in range(user_max):
                    row_index = tmp_user_row_index + user_index + 1
                    ws.cell(row=row_index, column=real_table_info['receiver__username']['col_index'], value=user)

                # 执行人列行数 在user填充后自加
                real_table_info['receiver__username']['cur_row_index'] += user_max
                for date_range, tasks in date_range_data.items():
                    # 当前人时间段的任务行数
                    tmp_cur_row_index = real_table_info[date_range]['cur_row_index']
                    # 上一用户没有包含当前时间段，则使用上一个用户的最后记录的行号
                    for task in tasks:
                        tmp_cur_row_index += 1
                        for key, val in task.items():
                            col_index = real_table_info[date_range]['sub'][key]
                            ws.cell(row=tmp_cur_row_index, column=col_index, value=val)
                # 当当前时间段的行索引对齐用户行
                cls.update_date_range_index(user_max, real_table_info)
            # 当前项目遍历结束后，项目索引对齐
            real_table_info['project']['cur_row_index'] += project_max

        # 项目最长的排前面
        # header_row_1 = ['项目', '执行人', '2024-08-01~2024-08-14', '', '', '', '', '2024-08-15~2024-08-21', '', '', '', '']  # 通过函数获取
        # header_row_2 = ['    ', '     ', 'ID', '任务名', '任务内容', '任务目标', '工作量', 'ID', '任务名', '任务内容', '任务目标', '工作量']  # 通过函数获取
        for merge in merge_list:
            ws.merge_cells(merge)

        return wb

    def get_project_max(cls, project_dict):
        ret_dict = {
            'project_max': 0,
            'users': {}
        }
        for user, date_range_data in project_dict.items():
            cur_user_max = 0
            for date_range, tasks in date_range_data.items():
                task_length = len(tasks)
                if task_length > cur_user_max:
                    cur_user_max = task_length
            ret_dict['users'][user] = cur_user_max
            ret_dict['project_max'] += cur_user_max
        return ret_dict

    def get(self, request, *args, **kwargs):
        # 假设你有一个函数来获取任务数据
        # 获取查询参数
        filter_params = request.query_params.dict()
        # 从查询参数中提取group（如果存在）
        group_id = filter_params.pop('group', None)

        # 创建FilterSet实例
        task_filter = TaskFilter(data=filter_params, queryset=Task.objects.all())

        # 检查过滤是否有效
        if not task_filter.is_valid():
            return Response(task_filter.errors, status=status.HTTP_400_BAD_REQUEST)

        filtered_tasks = task_filter.qs

        # 如果提供了group_id，则进一步过滤tasks
        if group_id:
            try:
                group = Group.objects.get(id=group_id)
                # 使用Django ORM来查询接收者在给定组内的任务
                filtered_tasks = filtered_tasks.filter(receiver__group__id=group_id)
            except Group.DoesNotExist:
                # 处理组不存在的情况
                return Response({"error": "Group not found"}, status=status.HTTP_404_NOT_FOUND)
                # 排序 id降序
        if filtered_tasks.count() == 0:
            return Response({"error": "not tasks"}, status=status.HTTP_404_NOT_FOUND)
        # 准备数据框
        time_range = {'min': "", 'max': ""}
        tasks = filtered_tasks.values('project', 'receiver__username', 'id', 'name', 'content', 'challenge', 'workload', 'start_time', 'deadline_time')
        wb = self.create_excel_with_data(tasks)

        # 创建一个输出字节流
        output = io.BytesIO()

        # 保存工作簿到输出字节流
        wb.save(output)

        # 获取输出字节流中的值
        output.seek(0)

        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        if isinstance(time_range['min'], datetime):
            time_range['min'] = time_range['min'].strftime("%Y-%m-%d")
        if isinstance(time_range['max'], datetime):
            time_range['max'] = time_range['max'].strftime("%Y-%m-%d")
        # excel_name = "tasks%s~%s.xlsx"%(time_range['min'],time_range['max'])
        excel_name = "export_tasks.xlsx"
        # content_disposition = 'attachment; filename="%s"'% excel_name
        content_disposition = 'attachment; filename="%s"' % excel_name
        response['Content-Disposition'] = content_disposition

        return response


class ImportExcel(APIView):
    def post(self, request, *args, **kwargs):
        # 从 request.FILES 获取文件
        userid = self.kwargs.get('userid')
        creator = None
        try:
            creator = User.objects.get(id=userid)
        except User.DoesNotExist:
            return Response({"error": "No creator !"}, status=400)  # 返回一个空的Job查询集

        excel_file = request.FILES['file']

        # 验证是否收到了文件
        if not excel_file:
            return Response({"error": "No file uploaded."}, status=400)

            # 验证文件类型（可选，但推荐）
        # if not isinstance(excel_file, InMemoryUploadedFile):
        #     return Response({"error": "Uploaded file is not an InMemoryUploadedFile."}, status=400)

        # 使用 pandas 读取 Excel 文件
        try:
            df1 = pd.read_excel(excel_file)
            df2 = pd.read_excel(excel_file, header=[0, 1])
        except ValueError:
            return Response({"error": "Uploaded file is not an Excel file."}, status=400)
        head1 = set(df1.columns)
        head2 = df2.columns
        head2_1 = [i[0].strip() for i in head2]
        head2_2 = [i[1].strip() for i in head2 if isinstance(i[1], str)]
        col_style1 = ['ID', '任务名', '执行人', '发布人', '创建人', '开始时间', '截止时间', '工作量/天', '项目', '关联任务', '当前状态', '任务内容', '挑战目标', '反馈信息']
        if set(col_style1).issubset(head1):
            for index, row in df1.iterrows():
                try:
                    task = Task.objects.get(id=row['ID'])
                    # 暂时不能通过excel改变数据库已存在的任务
                except:
                    task = Task()
                receiver, publisher = None, None
                try:
                    receiver = User.objects.get(username=row['执行人'])
                except User.DoesNotExist:
                    continue
                try:
                    publisher = User.objects.get(username=row['发布人'])
                except:
                    pass
                task.name = row['任务名']
                if receiver:
                    task.receiver = receiver
                if publisher:
                    task.publisher = publisher
                if creator:
                    task.creator = creator
                time_format1 = "%Y-%m-%d"
                try:
                    if isinstance(row['开始时间'], datetime):
                        task.start_time = row['开始时间'].date()
                    else:
                        task.start_time = datetime.strptime(row['开始时间'], time_format1)
                except TypeError:
                    task.start_time = row['开始时间'].to_pydatetime()
                try:
                    if isinstance(row['截止时间'], datetime):
                        task.deadline_time = row['截止时间'].date()
                    else:
                        task.deadline_time = datetime.strptime(row['截止时间'], time_format1)
                except TypeError:
                    task.deadline_time = row['截止时间'].to_pydatetime()
                task.workload = row['工作量/天']
                task.project = row['项目']
                task.content = row['任务内容']
                task.challenge = row['挑战目标']
                task.feedback = row['反馈信息']
                task.status = TaskStatus.objects.get(name=common.TASK_STATUS_DRAFT)
                task.save()

        elif '任务目标' in head2_2 and '工作量' in head2_2 and ('执行人' in head2_1 or '测试人员' in head2_1):
            data_map = {'receiver': -1, 'creator': -1, 'date': {}}
            # {'receiver': 4, 'date': {'7.29-8.4': {'任务内容': 6, '任务目标': 7, '工作量': 8}, '8.5-8.11': {'任务内容': 9, '任务目标': 10, '工作量': 11}}}
            user_map = {}
            for index in range(len(head2_1)):
                head_item = head2_1[index]
                if head_item == '执行人' or head_item == '测试人员':
                    data_map['receiver'] = index
                if head_item == '测试负责人':
                    data_map['creator'] = index
                if head_item == '项目名称' or head_item == '项目':
                    data_map['project'] = index
                elif is_date_range(head_item):
                    second_item = head2_2[index]
                    if second_item not in ['任务内容', '任务目标', '工作量', '质量目标']:
                        continue
                        # return Response({"error": ('%s not support' % second_item)}, status=400)
                    if head_item not in data_map['date']:
                        data_map['date'][head_item] = {}
                    data_map['date'][head_item][second_item] = index
            for index, row in df2.iterrows():
                values = list(row.values)
                if isinstance(values[0], str) and "汇总" in values[0]:
                    continue
                tmp_reciver = values[data_map['receiver']]
                tmp_creator = values[data_map['creator']]
                if isinstance(tmp_reciver, float) and math.isnan(tmp_reciver):
                    continue
                project = values[data_map['project']]
                for date_range, date_map in data_map['date'].items():
                    # TODO 处理用户不存在
                    flag, start, end = parse_date_range(date_range)
                    if not flag:
                        continue
                    task = Task()
                    try:
                        receiver = User.objects.get(username__icontains=tmp_reciver)
                    except:
                        continue
                    try:
                        if "/" in tmp_creator:
                            creator = User.objects.get(username__icontains=tmp_creator.split("/")[0])
                        else:
                            creator = User.objects.get(username__icontains=tmp_creator)
                    except:
                        creator = None
                    task.start_time = start
                    task.deadline_time = end
                    task.receiver = receiver
                    if creator:
                        task.creator = creator
                    task.project = project
                    task_name, task_content, task_challenge = '', '', ''
                    if "任务内容" in date_map:
                        task_content = values[date_map['任务内容']]
                        if task_content:
                           try:
                               task_name = task_content[:10]
                               split_list = task_content.split("##")
                               if(len(split_list) == 2):
                                   task_name = split_list[0]
                                   task_content = split_list[1]
                           except:
                               print('error name:%s'% task_content)
                               task_name = '-'    
                    if "任务目标" in date_map and not task_content and isinstance(values[date_map['任务目标']], str):
                        task_content = values[date_map['任务目标']]
                        task_name = task_content[:10]
                    if "任务目标" in date_map and not task_challenge and isinstance(values[date_map['任务目标']], str):
                        task_challenge = values[date_map['任务目标']]
                    if "质量目标" in date_map and not task_challenge and isinstance(values[date_map['质量目标']], str):
                        task_challenge = values[date_map['质量目标']]
                    task.name = task_name if not pd.isna(task_name) else ""
                    task.content = task_content if not pd.isna(task_content) else ""
                    task.challenge = task_challenge if not pd.isna(task_challenge) else ""
                    if pd.isna(values[date_map['工作量']]) or str(values[date_map['工作量']]) == '-':
                        task.workload = 0
                    else:
                        task.workload = values[date_map['工作量']]
                    task.status = TaskStatus.objects.get(name=common.TASK_STATUS_DRAFT)
                    task.save()
        else:
            return Response({"error": "Not support!"}, status=400)
            pass
        # 返回响应
        return Response({"message": "Excel file processed successfully", "data": {}}, status=200)


class ExportTemplateExcel(APIView):
    def get(self, request, *args, **kwargs):
        # 指定要下载的 Excel 文件路径
        file_path = os.path.join(settings.BASE_DIR, 'cfg/任务导入模板.xlsx')

        # 检查文件是否存在
        if not default_storage.exists(file_path):
            return HttpResponse("File not found", status=404)

        # 使用 FileResponse 返回文件
        response = FileResponse(open(file_path, 'rb'))
        response['Content-Disposition'] = 'attachment; filename=task_template.xlsx'
        return response
