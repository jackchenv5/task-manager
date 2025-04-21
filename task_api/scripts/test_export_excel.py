import os
import string
import sys
from datetime import datetime
# 启动django
from pathlib import Path

import django
from openpyxl import Workbook

from openpyxl.styles import Font, Color, PatternFill, Border, Side, Alignment

# Build paths inside the project like this: BASE_DIR / 'subdir'.
BASE_DIR = str(Path(__file__).resolve().parent.parent)
sys.path.insert(0, BASE_DIR)

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'task_api.settings')
django.setup()

from django.contrib.auth import get_user_model  # 导入Django的用户模型
from task.models import Task

User = get_user_model()



def get_col_letter(index):
    uppercase_letters = list(string.ascii_uppercase)
    new_index =  index%len(uppercase_letters)
    count =  int(index/len(uppercase_letters))
    COL_LETTER = uppercase_letters[new_index]
    if count > 0:
        AFFIX = uppercase_letters[count-1]
        COL_LETTER  = AFFIX + COL_LETTER
    return COL_LETTER

def get_data(tasks):
    date_range_list = []
    data_dict = {}
    for task in tasks:
        start_date = task['start_time'].strftime("%Y-%m-%d")
        deadline_date = task['deadline_time'].strftime("%Y-%m-%d")
        task['start_time'] = start_date
        task['deadline_time'] = deadline_date
        project_name = task['project']
        receiver__username = task['receiver__username']
        date_range = (start_date, deadline_date)
        if date_range not in date_range_list:
            date_range_list.append(date_range)
        if project_name not in data_dict:
            data_dict[project_name] = {}
        if receiver__username not in data_dict[project_name]:
            data_dict[project_name][receiver__username] = {}
        if date_range not in data_dict[project_name][receiver__username]:
            data_dict[project_name][receiver__username][date_range] = []
        del task['project']
        del task['receiver__username']
        del task['start_time']
        del task['deadline_time']
        data_dict[project_name][receiver__username][date_range].append(task)
    return data_dict, date_range_list

def update_date_range_index(row_num,ret_dict):
    for key,data in ret_dict.items():
        if key in ['project','receiver__username']:
            continue
        else:
            ret_dict[key]['cur_row_index'] += row_num

def get_table_info(data_list):
    ret_dict = {
        'project': {'col_index': 1, 'cur_row_index': 2, 'name': '项目', 'width': 30},
        'receiver__username': {'col_index': 2, 'cur_row_index': 2, 'name': '执行者', 'width': 10},
        # ('2024-08-01', '2024-08-14'): {'col_index': 3, 'cur_row': 2, 'sub': {'id': 3, 'name': 4, 'content': 5, 'challenge': 6, 'workload': 7}},
        # ('2024-08-15', '2024-08-29'): {'col_index': 8, 'cur_row': 2, 'sub': {'id': 8, 'name':9, 'content': 10, 'challenge': 11, 'workload': 12}}
    }
    width_config = {'A':50,'B':10}
    header_row_1 = ['项目', '执行人']  # 通过函数获取
    header_row_2 = ['', '']  # 通过函数获取
    merge_list = ['A1:A2','B1:B2']

    template_dict = {'id': 3, 'name': 4, 'content': 5, 'challenge': 6, 'workload': 7}
    template_width_dict = {'id': 10, 'name': 30, 'content': 50, 'challenge': 50, 'workload': 10}
    template_dict_len = len(template_dict)
    name_list = ['ID', '任务名', '任务内容', '任务目标', '工作量']
    template_dict_list = ['' for i in range(template_dict_len-1)]
    col_index_base = 3
    cur_row_index_base = 2
    len_dict = len(template_dict)
    sorted_dates = sorted(data_list, key=lambda d: datetime.strptime(d[0], "%Y-%m-%d"))  # key参数实际上是可选的，因为datetime对象已经可比较
    date_range_count = 0
    letter_index = 2
    for date_range in sorted_dates:
        header_row_1 += ['%s~%s'%(date_range[0],date_range[1])] + template_dict_list
        start_letter = get_col_letter(len(header_row_2))
        end_letter = get_col_letter(len(header_row_2)+template_dict_len-1)
        merge_list.append('%s1:%s1'%(start_letter,end_letter))
        header_row_2 += name_list
        ret_dict[date_range]={'col_index':-1,'cur_row_index':-1,'sub':{}}
        ret_dict[date_range]['col_index'] = col_index_base + len_dict * date_range_count
        ret_dict[date_range]['cur_row_index'] = cur_row_index_base
        for key, value in template_dict.items():
            print(get_col_letter(letter_index),key)
            width_config[get_col_letter(letter_index)] = template_width_dict[key]
            ret_dict[date_range]['sub'][key] = template_dict[key] + len_dict * date_range_count
            letter_index += 1
        date_range_count += 1
    print(width_config)
    return ret_dict,header_row_1,header_row_2,merge_list,width_config


def create_excel_with_data(tasks):
    # 创建一个新的Excel工作簿
    wb = Workbook()
    ws = wb.active

    real_data, date_range_list = get_data(tasks)
    real_table_info,header_row_1,header_row_2,merge_list,width_config = get_table_info(date_range_list)
    for key,value in width_config.items():
        ws.column_dimensions[key].width = value
    # 项目最长的排前面
    # header_row_1 = ['项目', '执行人', '2024-08-01~2024-08-14', '', '', '', '', '2024-08-15~2024-08-21', '', '', '', '']  # 通过函数获取
    # header_row_2 = ['    ', '     ', 'ID', '任务名', '任务内容', '任务目标', '工作量', 'ID', '任务名', '任务内容', '任务目标', '工作量']  # 通过函数获取

    # 写入表头
    # 字体样式：加粗、大小为12
    bold_font = Font(bold=True, size=14,color="FFFFFF")
    font2 =  Font(bold=True, size=12,color="FFFFFF")
    # 填充样式：浅灰色背景
    fill1 = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    fill2 = PatternFill(start_color='666666', end_color='666666', fill_type='solid')

    # 边框样式：细边框
    thin_border = Side(border_style="thin", color="000000")
    border = Border(top=thin_border, left=thin_border, right=thin_border, bottom=thin_border)
    header_alignment = Alignment(horizontal='center', vertical='center')  # 水平和垂直居中

    for col_num, header in enumerate(header_row_1, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = bold_font
        cell.fill = fill1
        cell.border = border
        cell.alignment = header_alignment
    for col_num, header in enumerate(header_row_2, 1):
        cell = ws.cell(row=2, column=col_num, value=header)
        cell.font = font2
        cell.fill = fill2
        cell.border = border
        cell.alignment = header_alignment
    # 获取用户在各个时间段的任务数，获取最大任务 project_user_max_len
    for project, data in real_data.items():
        max_info = get_project_max(data)
        project_max = max_info['project_max']
        # 填充项目列
        cur_project_row_index = real_table_info['project']['cur_row_index']
        for project_index in range(project_max):
            cur_project_row_index += 1
            # 是否需要合并
            # if row_index == 1:
            #     ws.cell(row=row_index, column=table_info['project']['col_index'],value=project)
            ws.cell(row=cur_project_row_index, column=real_table_info['project']['col_index'], value=project)
        # 合并项目列
        # 填充用户

        for user, date_range_data in data.items():
            user_max = max_info['users'][user]
            tmp_user_row_index = real_table_info['receiver__username']['cur_row_index']

            for user_index in range(user_max):
                row_index = tmp_user_row_index + user_index + 1
                ws.cell(row=row_index, column=real_table_info['receiver__username']['col_index'], value=user)

            # 执行人列行数 在user填充后自加
            real_table_info['receiver__username']['cur_row_index'] += user_max
            for date_range, tasks in date_range_data.items():
                # 当前人时间段的任务行数
                tmp_cur_row_index = real_table_info[date_range]['cur_row_index']
                # 上一用户没有包含当前时间段，则使用上一个用户的最后记录的行号
                for task in tasks:
                    tmp_cur_row_index += 1
                    for key, val in task.items():
                        col_index = real_table_info[date_range]['sub'][key]
                        ws.cell(row=tmp_cur_row_index, column=col_index, value=val)
            # 当当前时间段的行索引对齐用户行
            update_date_range_index(user_max,real_table_info)
        # 当前项目遍历结束后，项目索引对齐
        real_table_info['project']['cur_row_index'] += project_max

    # 项目最长的排前面
    # header_row_1 = ['项目', '执行人', '2024-08-01~2024-08-14', '', '', '', '', '2024-08-15~2024-08-21', '', '', '', '']  # 通过函数获取
    # header_row_2 = ['    ', '     ', 'ID', '任务名', '任务内容', '任务目标', '工作量', 'ID', '任务名', '任务内容', '任务目标', '工作量']  # 通过函数获取
    for merge in merge_list:
        ws.merge_cells(merge)
    # ws.merge_cells('A1:A2')
    # ws.merge_cells('B1:B2')
    # ws.merge_cells('C1:G1')
    # ws.merge_cells('H1:L1')

    # 设置列宽
    # ws.column_dimensions['A'].width = 10
    # ws.column_dimensions['B'].width = 20
    # ws.column_dimensions['C'].width = 10
    # ws.column_dimensions['D'].width = 30
    # ws.column_dimensions['E'].width = 30
    # ws.column_dimensions['F'].width = 30
    # ws.column_dimensions['G'].width = 10
    # ws.column_dimensions['H'].width = 10
    # ws.column_dimensions['I'].width = 30
    # ws.column_dimensions['J'].width = 30
    # ws.column_dimensions['K'].width = 30
    # ws.column_dimensions['K'].width = 10

    # 保存Excel文件
    filename = f"Tasks_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    print(f"Excel file '{filename}' has been created.")


def get_project_max(project_dict):
    ret_dict = {
        'project_max': 0,
        'users': {}
    }
    for user, date_range_data in project_dict.items():
        cur_user_max = 0
        for date_range, tasks in date_range_data.items():
            task_length = len(tasks)
            if task_length > cur_user_max:
                cur_user_max = task_length
        ret_dict['users'][user] = cur_user_max
        ret_dict['project_max'] += cur_user_max
    return ret_dict


if __name__ == '__main__':
    task_queryset = Task.objects.filter(project='RT-V008R012C000(PR072)_信创核心路由器产品开发')[:100]
    tasks = task_queryset.values('project', 'receiver__username', 'id', 'name', 'content', 'challenge', 'workload', 'start_time', 'deadline_time')
    create_excel_with_data(tasks)
